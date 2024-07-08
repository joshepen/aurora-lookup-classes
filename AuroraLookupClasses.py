from bs4 import BeautifulSoup
from auroranav.AuroraNav import AuroraNav
import os
import openpyxl as xl
import threading

INPUT_DIR = "input/"
OUTPUT_DIR = "output/"


def main():
    i = 1
    files = os.listdir(INPUT_DIR)
    filename = "courses%d.txt" % i
    while filename in files:
        LookupClasses(INPUT_DIR + filename)
        i += 1
        filename = "courses%d.txt" % i


def LookupClasses(filename):
    semester = GetSemester(filename)
    names = GetClassNames(filename)
    userID = GetUserID(filename)
    password = GetPassword(filename)

    nav = AuroraNav(headless=False)
    nav.OpenAurora()
    nav.Login(userID, password)
    data = []
    for name in names:
        data.append(GetClassInfo(nav, semester, name))

    CourseInfoToSpreadsheet(OUTPUT_DIR + semester + ".xlsx", data)

    nav.CloseWindow()


def GetUserID(filepath):
    userIDIndex = 0
    with open(filepath, "r") as file:
        return file.readlines()[userIDIndex].strip()


def GetPassword(filepath):
    passwordIndex = 1
    with open(filepath, "r") as file:
        return file.readlines()[passwordIndex].strip()


def GetCatalogInfo(nav: AuroraNav, semester, courseName):
    nav.HomePage()
    nav.GoToPage("Enrolment & Academic Records")
    nav.GoToPage("Registration and Exams")
    nav.GoToPage("Course Catalog")

    nav.SelectTerm(semester)
    nav.SelectDepartment(courseName.split(" ")[0])
    if nav.GoToPage(courseName):
        return ProcessCatalogPage(
            BeautifulSoup(nav.GetPageContents(), features="html.parser")
        )
    return None


def GetLUCInfo(nav: AuroraNav, semester, name):
    nav.HomePage()
    nav.GoToPage("Enrolment & Academic Records")
    nav.GoToPage("Registration and Exams")
    nav.GoToPage("Look Up Classes")

    nav.SelectTerm(semester)
    nav.SelectDepartment(name.split(" ")[0])
    if nav.GoToLookupClass(name.split(" ")[1]):
        content = BeautifulSoup(nav.GetPageContents(), features="html.parser")
        content = ProcessLUCPage(content)
        return content
    return []


def GetSemester(filename):
    semesterIndex = 2
    with open(filename, "r") as file:
        return file.readlines()[semesterIndex].strip()


def GetClassNames(filename):
    classesIndex = 3
    names = []
    with open(filename, "r") as file:
        for line in file.readlines()[classesIndex:]:
            names.append(line.strip())
    return names


def ProcessLUCPage(pageContents: BeautifulSoup):
    table = pageContents.find("table", class_="datadisplaytable").tbody

    data = []
    wantedColumns = GetWantedColumnIndices(table.contents[2])

    currItem = {}
    for name in wantedColumns:
        currItem[name] = []

    for trIndex in range(3, len(table.contents)):
        if (
            table.contents[trIndex] != "\n"
            and table.contents[trIndex].find("b") == None
            and table.contents[trIndex].find("hr") == None
        ):
            for columnName in wantedColumns:
                columnIndex = wantedColumns[columnName]
                value = table.contents[trIndex].contents[columnIndex].string
                currItem[columnName].append(value)

        elif table.contents[trIndex].find("hr") not in [None, -1]:
            data.append(currItem)
            currItem = {}
            for name in wantedColumns:
                currItem[name] = []

    return data


def ProcessCatalogPage(page: BeautifulSoup):
    return page.find("td", class_="ntdefault").get_text()


def GetWantedColumnIndices(headerRow: BeautifulSoup):
    wantedColumns = [
        "CRN,",
        "Sec",
        "Cmp",
        "Title",
        "Days",
        "Time",
        "Instructor",
        "Date",
        "Rem",
    ]
    wantedIndices = {}
    for i in range(len(headerRow.contents)):
        for name in wantedColumns:
            if name in headerRow.contents[i].get_text():
                wantedIndices[headerRow.contents[i].get_text()] = i

    return wantedIndices


def GetClassInfo(nav: AuroraNav, semester, name):
    data = {}
    data["table"] = GetLUCInfo(nav, semester, name)
    data["description"] = GetCatalogInfo(nav, semester, name)
    data["name"] = name
    return data


def CourseInfoToSpreadsheet(filepath, data: list[dict]):
    wb = xl.Workbook()
    sortedData = sorted(data, key=CourseDataSortFn)
    for course in sortedData:
        ws = wb.create_sheet(course["name"])
        ws["A1"].value = course["description"]
        if len(course["table"]) > 0:
            rowIndex = 2
            colIndex = 1
            # Create header row
            for header in course["table"][0]:
                ws.cell(row=rowIndex, column=colIndex).value = header
                colIndex += 1
            rowIndex += 1
            # Add data per section (time)
            for section in course["table"]:
                colIndex = 1
                for header in section:
                    for i in range(len(section[header])):
                        ws.cell(row=rowIndex + i, column=colIndex).value = section[
                            header
                        ][i]
                    colIndex += 1
                rowIndex += 1
    wb.remove(wb["Sheet"])
    wb.save(filepath)


def CourseDataSortFn(course: dict):
    if len(course["table"]) == 0:
        return str(
            [
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
                0xFF,
            ]
        )
    else:
        return course["name"]


if __name__ == "__main__":
    main()
