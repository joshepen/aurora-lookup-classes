from objects.Course import Course
import openpyxl as xl
import itertools


class OutputExcel:
    @staticmethod
    def output(data: list[Course], filepath: str) -> any:
        wb = xl.Workbook()
        sortedData = sorted(data, key=OutputExcel.CourseDataSortFn)
        for course in sortedData:
            ws = wb.create_sheet(course.GetFullName())
            ws["A1"].value = course.GetDescription()
            if len(course) > 0:
                rowIndex = 2
                # Create header row
                headers = course.GetHeaders()
                for i in range(len(headers)):
                    ws.cell(row=rowIndex, column=i + 1).value = headers[i]
                rowIndex += 1
                # Add data per section
                rows = itertools.chain.from_iterable(course.GetSections())
                for row in rows:
                    for colIndex in range(len(row)):
                        ws.cell(row=rowIndex, column=colIndex + 1).value = row[colIndex]
                    rowIndex += 1
        wb.remove(wb["Sheet"])
        wb.save(filepath + ".xlsx")

    @staticmethod
    def CourseDataSortFn(course: Course):
        if len(course) == 0:
            # Empty courses should have been removed already but just in case
            return str(
                [
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                    0xFF,
                ]
            )
        else:
            return course.GetName()
